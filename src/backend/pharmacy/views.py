from rest_framework import viewsets, status, filters
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.db import transaction
import openpyxl

from .models import Medicine, StockMovement, PharmacyStock, GlobalSettings, MedicineBatch, MedicalCenter
from .serializers import (
    MedicineSerializer, StockMovementSerializer, PharmacyStockSerializer,
    GlobalSettingsSerializer, MedicineBatchSerializer, MedicalCenterSerializer
)

class MedicalCenterViewSet(viewsets.ModelViewSet):
    queryset = MedicalCenter.objects.all()
    serializer_class = MedicalCenterSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name']
    ordering_fields = ['name']
    pagination_class = None

class GlobalSettingsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        settings = GlobalSettings.load()
        serializer = GlobalSettingsSerializer(settings)
        return Response(serializer.data)

    def put(self, request):
        settings = GlobalSettings.load()
        serializer = GlobalSettingsSerializer(settings, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class MedicineViewSet(viewsets.ModelViewSet):
    queryset = Medicine.objects.all()
    serializer_class = MedicineSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'reference_number']
    ordering_fields = ['name']
    pagination_class = None

class StockMovementViewSet(viewsets.ModelViewSet):
    queryset = StockMovement.objects.all()
    serializer_class = StockMovementSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['medical_center', 'movement_type', 'medicine']
    ordering_fields = ['date', 'created_at']

class PharmacyStockViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PharmacyStock.objects.all()
    serializer_class = PharmacyStockSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['medical_center', 'medicine']
    ordering_fields = ['quantity']
    pagination_class = None

class MedicineBatchViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = MedicineBatch.objects.filter(quantity__gt=0)
    serializer_class = MedicineBatchSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['medical_center', 'medicine', 'expiration_date']
    ordering_fields = ['expiration_date']
    pagination_class = None

class ExcelRequisitionImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, site_id, *args, **kwargs):
        if 'file' not in request.FILES:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            site = MedicalCenter.objects.get(id=site_id)
        except MedicalCenter.DoesNotExist:
            return Response({"detail": "Invalid medical center ID."}, status=status.HTTP_400_BAD_REQUEST)

        file_obj = request.FILES['file']
        if not file_obj.name.endswith(('.xlsx', '.xls')):
            return Response({"detail": "Invalid file format. Only .xlsx and .xls are supported."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
            
            # Expected headers: Medicine Name, Reference Number, Unit, Quantity, Expiration Date (YYYY-MM-DD)
            # Find column indices
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip().lower() if h else '' for h in header_row]
            
            try:
                name_idx = next(i for i, h in enumerate(headers) if 'name' in h or 'nom' in h)
                qty_idx = next(i for i, h in enumerate(headers) if 'quantit' in h or 'qty' in h)
            except StopIteration:
                return Response({"detail": "Missing required columns: Medicine Name or Quantity."}, status=status.HTTP_400_BAD_REQUEST)
                
            ref_idx = next((i for i, h in enumerate(headers) if 'ref' in h), None)
            unit_idx = next((i for i, h in enumerate(headers) if 'unit' in h), None)
            exp_idx = next((i for i, h in enumerate(headers) if 'exp' in h or 'date' in h), None)

            movements_created = 0
            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[name_idx] or not row[qty_idx]:
                        continue
                        
                    med_name = str(row[name_idx]).strip()
                    try:
                        quantity = int(row[qty_idx])
                    except (ValueError, TypeError):
                        continue # Skip invalid quantity
                        
                    if quantity <= 0:
                        continue
                        
                    ref = str(row[ref_idx]).strip() if ref_idx is not None and row[ref_idx] else None
                    unit = str(row[unit_idx]).strip() if unit_idx is not None and row[unit_idx] else 'Unité'
                    exp_date = row[exp_idx] if exp_idx is not None else None
                    if hasattr(exp_date, 'date'):
                        exp_date = exp_date.date()

                    # Get or Create Medicine
                    medicine, created = Medicine.objects.get_or_create(
                        name=med_name,
                        defaults={
                            'reference_number': ref,
                            'unit': unit
                        }
                    )

                    # Create Stock Movement (IN)
                    StockMovement.objects.create(
                        medical_center=site,
                        medicine=medicine,
                        movement_type='IN',
                        quantity=quantity,
                        expiration_date=exp_date,
                        notes='Imported from Excel Requisition'
                    )
                    movements_created += 1

            return Response({"success": True, "movements_created": movements_created}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"detail": f"Error processing file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

class ExcelConsumptionImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, site_id, *args, **kwargs):
        if 'file' not in request.FILES:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            site = MedicalCenter.objects.get(id=site_id)
        except MedicalCenter.DoesNotExist:
            return Response({"detail": "Invalid medical center ID."}, status=status.HTTP_400_BAD_REQUEST)

        file_obj = request.FILES['file']
        if not file_obj.name.endswith(('.xlsx', '.xls')):
            return Response({"detail": "Invalid file format. Only .xlsx and .xls are supported."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
            
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip().lower() if h else '' for h in header_row]
            
            try:
                name_idx = next(i for i, h in enumerate(headers) if 'name' in h or 'nom' in h)
                qty_idx = next(i for i, h in enumerate(headers) if 'quantit' in h or 'qty' in h)
            except StopIteration:
                return Response({"detail": "Missing required columns: Medicine Name or Quantity."}, status=status.HTTP_400_BAD_REQUEST)

            movements_created = 0
            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[name_idx] or not row[qty_idx]:
                        continue
                        
                    med_name = str(row[name_idx]).strip()
                    try:
                        quantity = int(row[qty_idx])
                    except (ValueError, TypeError):
                        continue
                        
                    if quantity <= 0:
                        continue

                    # For OUT movement, we expect the medicine to exist
                    try:
                        medicine = Medicine.objects.get(name__iexact=med_name)
                    except Medicine.DoesNotExist:
                        continue # Skip if medicine not found

                    # Create Stock Movement (OUT)
                    StockMovement.objects.create(
                        medical_center=site,
                        medicine=medicine,
                        movement_type='OUT',
                        quantity=quantity,
                        notes='Imported from Excel Consumption'
                    )
                    movements_created += 1

            return Response({"success": True, "movements_created": movements_created}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"detail": f"Error processing file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


